from openpyxl import load_workbook
import mysql.connector

criar_banco ='''CREATE TABLE teste1 (
                                 PLACA_ANTIGA varchar(200),
                                 CMT varchar(200),
                                 NR_PASSAGEIROS varchar(200),
                                 BUSCA varchar(200)); '''

                                 
def main():
        # Caso não tenha configurado o banco ainda, crie um banco com nome de infocar e  
        # descomente a linha abaixo para gerar a tabela:
        # executa_no_banco(criar_banco)
        # OBS: Cheque o metodo executa_no_banco para conferir os dados de conexão

        dados_para_o_banco = extrai_info_do_excel('teste-ti-processamento_fornecedor_fora_do_ar.xlsx')
        if dados_para_o_banco:
                comando_sql=""
                for i in dados_para_o_banco:
                        comando_sql = f'''
                                INSERT INTO teste1 (PLACA_ANTIGA,CMT,NR_PASSAGEIROS,BUSCA) VALUES (
                                "{i["PLACA_ANTIGA"]}",
                                "{i["CMT"]}",
                                "{i["NR_PASSAGEIROS"]}",
                                "{i["BUSCA"]}");
                        '''
                        executa_no_banco(comando_sql)
                print('itens adicionados ao banco de dados com sucesso')
                      
        else: print("Erro ao extrair dados do Excel")


def extrai_info_do_excel(nome_arquivo):
        try:
                arquivo = load_workbook(filename = nome_arquivo)
                planilha = arquivo['planilha']

                registros = []

                for linha in planilha.iter_rows(min_row=2, values_only=True):
                        nova_linha = {"PLACA_ANTIGA":linha[0],
                                        "CMT":linha[1],
                                        "NR_PASSAGEIROS":linha[2],
                                        "BUSCA":linha[3]}
                                        
                        registros.append(nova_linha)
                return registros
        except Exception as erro:
                print("Algo deu errado ao executar o 'método extrai_info_do_excel'\n\n", erro)
                return False

def executa_no_banco(sql):
        try:
        
                conexao = mysql.connector.connect(host='localhost', database='infocar',user='root',password='')
                cur = conexao.cursor()
                cur.execute(sql)
                conexao.commit()
                cur.close()
                conexao.close()

        except Exception as erro:
                print("Algo deu errado ao executar método 'executa_no_banco'\n\n",erro)
                return False


if __name__ == "__main__": main()


